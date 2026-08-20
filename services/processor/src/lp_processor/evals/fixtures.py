from __future__ import annotations

import io
import time
from dataclasses import dataclass

from openpyxl import Workbook

from lp_processor.extractors.heuristic import HeuristicExtractor
from lp_processor.parsers.base import ParserNotWiredError
from lp_processor.parsers.routing import select_parser
from lp_processor.routing_policy import decide_route


def _pdf(text: str | None) -> bytes:
    if text is None:
        contents = b"<< /Length 0 >>\nstream\nendstream"
    else:
        stream = f"BT /F1 12 Tf 36 720 Td ({text}) Tj ET".encode("latin-1")
        contents = b"<< /Length %d >>\nstream\n" % len(stream) + stream + b"\nendstream"
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>",
        contents,
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    out = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, body in enumerate(objects, start=1):
        offsets.append(len(out))
        out.extend(f"{index} 0 obj\n".encode("ascii"))
        out.extend(body)
        out.extend(b"\nendobj\n")
    xref_start = len(out)
    out.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    out.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        out.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    out.extend(
        f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF\n".encode("ascii")
    )
    return bytes(out)


def pricing_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Pricing"
    ws["A1"] = "Labor category"
    ws["B1"] = "Unit price"
    ws["A2"] = "Armed officer"
    ws["B2"] = 38.25
    ws["B2"].number_format = '"$"#,##0.00'
    ws["C2"] = "=B2*2080"
    ws["A3"] = "Due"
    ws["B3"] = "2026-03-15"
    ws.merge_cells("A4:B4")
    ws["A4"] = "Merged overtime note"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


DIGITAL_RFP_TEXT = (
    "Northside ISD RFP 26-04 Provide 24/7 coverage. "
    "Due date 2026-03-15. Client Northside ISD. "
    "Requirement: Maintain current insurance certificates."
)


@dataclass(frozen=True)
class FixtureCase:
    case_id: str
    package_role: str
    filename: str
    mime_type: str
    payload: bytes
    gold_cells: dict[str, str]
    gold_requirements: list[str]
    gold_entities: list[str]
    gold_dates: list[str]
    expect_escalate: bool
    expected_parser_id: str
    form_fields: list[str]


def fixture_cases() -> list[FixtureCase]:
    return [
        FixtureCase(
            case_id="xlsx_pricing_workbook",
            package_role="proposal_pricing",
            filename="pricing.xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            payload=pricing_workbook(),
            gold_cells={"Pricing!B2": "38.25", "Pricing!A2": "Armed officer"},
            gold_requirements=[],
            gold_entities=[],
            gold_dates=["2026-03-15"],
            expect_escalate=False,
            expected_parser_id="xlsx-openpyxl",
            form_fields=[],
        ),
        FixtureCase(
            case_id="digital_rfp_pdf",
            package_role="solicitation",
            filename="rfp.pdf",
            mime_type="application/pdf",
            payload=_pdf(DIGITAL_RFP_TEXT),
            gold_cells={},
            gold_requirements=["Provide 24/7 coverage", "Maintain current insurance certificates"],
            gold_entities=["Northside ISD"],
            gold_dates=["2026-03-15"],
            expect_escalate=False,
            expected_parser_id="pdf-native",
            form_fields=[],
        ),
        FixtureCase(
            case_id="scanned_pdf_empty_text",
            package_role="scanned_contract",
            filename="scan.pdf",
            mime_type="application/pdf",
            payload=_pdf(None),
            gold_cells={},
            gold_requirements=[],
            gold_entities=[],
            gold_dates=[],
            expect_escalate=True,
            expected_parser_id="ocr-mistral",
            form_fields=[],
        ),
        FixtureCase(
            case_id="docx_proposal",
            package_role="proposal_narrative",
            filename="proposal.docx",
            mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            payload=b"PK\x03\x04not-a-real-docx",
            gold_cells={},
            gold_requirements=[],
            gold_entities=[],
            gold_dates=[],
            expect_escalate=True,
            expected_parser_id="docx-native",
            form_fields=[],
        ),
        FixtureCase(
            case_id="form_checkbox_unwired",
            package_role="scorecard_form",
            filename="scorecard.pdf",
            mime_type="application/pdf",
            payload=_pdf(
                "Evaluator scorecard for Northside ISD. Checkbox fields are not parsed as form widgets in Phase 6. "
                "Yes No ratings remain page text until a form adapter is justified."
            ),
            gold_cells={},
            gold_requirements=[],
            gold_entities=[],
            gold_dates=[],
            expect_escalate=False,
            expected_parser_id="pdf-native",
            form_fields=["checkbox"],
        ),
    ]


def run_case(case: FixtureCase) -> dict:
    started = time.perf_counter()
    decision = decide_route(case.mime_type, case.filename, case.payload)
    escalated = False
    parser_id = decision.parser_id
    parsed = None
    drafts: list = []
    error = None
    try:
        parser = select_parser(case.mime_type, case.filename, case.payload)
        parsed = parser.parse(case.payload, mime_type=case.mime_type, filename=case.filename)
        drafts = HeuristicExtractor().extract(parsed)
        parser_id = parser.parser_id
    except ParserNotWiredError as exc:
        escalated = True
        error = str(exc)
    elapsed_ms = round((time.perf_counter() - started) * 1000, 2)

    cell_hits = 0
    for field, expected in case.gold_cells.items():
        match = next((draft for draft in drafts if draft.field == field), None)
        haystack = (match.normalized_value or match.raw_value or "") if match else ""
        if expected in haystack:
            cell_hits += 1
    cell_accuracy = 1.0 if not case.gold_cells else cell_hits / len(case.gold_cells)

    blob_parts = [(draft.normalized_value or draft.raw_value or "") for draft in drafts]
    if parsed:
        blob_parts.extend(page.text for page in parsed.pages)
    blob = " ".join(blob_parts)

    req_hits = sum(1 for req in case.gold_requirements if req in blob)
    requirement_recall = 1.0 if not case.gold_requirements else req_hits / len(case.gold_requirements)
    entity_hits = sum(1 for ent in case.gold_entities if ent in blob)
    entity_recall = 1.0 if not case.gold_entities else entity_hits / len(case.gold_entities)
    date_hits = sum(1 for date in case.gold_dates if date in blob)
    date_recall = 1.0 if not case.gold_dates else date_hits / len(case.gold_dates)

    provenance_ok = True
    if drafts and not case.expect_escalate:
        provenance_ok = all(
            (draft.source_section is not None) or (draft.source_page is not None) for draft in drafts
        )

    form_score: float | None = None if case.form_fields else 1.0
    scan_quality = "n/a"
    if case.case_id.startswith("scanned"):
        scan_quality = "escalated" if escalated else "missed_scan"

    route_ok = decision.parser_id == case.expected_parser_id and escalated == case.expect_escalate

    return {
        "case_id": case.case_id,
        "package_role": case.package_role,
        "parser_id": parser_id,
        "document_class": decision.document_class,
        "route_ok": route_ok,
        "escalated": escalated,
        "table_cell_accuracy": round(cell_accuracy, 3),
        "requirement_recall": round(requirement_recall, 3),
        "entity_recall": round(entity_recall, 3),
        "date_recall": round(date_recall, 3),
        "provenance_ok": provenance_ok,
        "forms": form_score,
        "scan_quality": scan_quality,
        "time_ms": elapsed_ms,
        "api_cost_usd": 0.0,
        "compute_cost_usd": 0.0,
        "error": error,
        "policy_version": decision.policy_version,
    }
