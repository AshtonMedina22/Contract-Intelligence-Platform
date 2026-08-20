from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from lp_processor.models import CellValue, NormalizedDocument, SheetStructure
from lp_processor.parsers.base import DocumentParser

XLSX_MIME = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroEnabled.12",
    "application/vnd.ms-excel",
}


def _stringify(value: Any) -> str | None:
    if value is None:
        return None
    return str(value)


class XlsxParser(DocumentParser):
    parser_id = "xlsx-openpyxl"

    def supports(self, mime_type: str | None, filename: str | None) -> bool:
        name = (filename or "").lower()
        if name.endswith((".xlsx", ".xlsm", ".xls")):
            return True
        return (mime_type or "") in XLSX_MIME

    def parse(self, payload: bytes, *, mime_type: str | None, filename: str | None) -> NormalizedDocument:
        formula_wb = load_workbook(io.BytesIO(payload), data_only=False, read_only=False)
        cached_wb = load_workbook(io.BytesIO(payload), data_only=True, read_only=False)
        warnings: list[str] = []
        sheets: list[SheetStructure] = []

        for index, sheet_name in enumerate(formula_wb.sheetnames):
            ws = formula_wb[sheet_name]
            cached_ws = cached_wb[sheet_name] if sheet_name in cached_wb.sheetnames else None
            skip: set[tuple[int, int]] = set()
            merged_at: dict[tuple[int, int], str] = {}
            merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
            for rng in ws.merged_cells.ranges:
                merged_at[(rng.min_row, rng.min_col)] = str(rng)
                for row in range(rng.min_row, rng.max_row + 1):
                    for col in range(rng.min_col, rng.max_col + 1):
                        if (row, col) != (rng.min_row, rng.min_col):
                            skip.add((row, col))

            cells: list[CellValue] = []
            for row in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row or 1,
                min_col=1,
                max_col=ws.max_column or 1,
            ):
                for cell in row:
                    if not isinstance(cell, Cell):
                        continue
                    if (cell.row, cell.column) in skip:
                        continue
                    formula = None
                    display = _stringify(cell.value)
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula = cell.value
                    cached = None
                    if cached_ws is not None:
                        cached_cell = cached_ws.cell(cell.row, cell.column)
                        cached = _stringify(cached_cell.value)
                    if formula is None and display is None and cached is None:
                        continue
                    cells.append(
                        CellValue(
                            sheet=sheet_name,
                            coordinate=cell.coordinate or f"{get_column_letter(cell.column)}{cell.row}",
                            row=cell.row,
                            column=cell.column,
                            data_type=cell.data_type or "n",
                            number_format=cell.number_format if cell.number_format != "General" else None,
                            formula=formula,
                            cached_value=cached,
                            display_value=cached if formula else display,
                            merged_range=merged_at.get((cell.row, cell.column)),
                        )
                    )

            sheets.append(
                SheetStructure(
                    name=sheet_name,
                    index=index,
                    max_row=ws.max_row or 0,
                    max_column=ws.max_column or 0,
                    merged_ranges=merged_ranges,
                    cells=cells,
                )
            )

        if not sheets:
            warnings.append("Workbook contained no sheets.")

        return NormalizedDocument(
            parser_id=self.parser_id,
            mime_type=mime_type,
            filename=filename,
            page_count=None,
            sheet_count=len(sheets),
            sheets=sheets,
            pages=[],
            warnings=warnings,
        )
