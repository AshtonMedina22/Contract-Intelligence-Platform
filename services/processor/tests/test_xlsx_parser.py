import io

from openpyxl import Workbook

from lp_processor.extractors.heuristic import HeuristicExtractor
from lp_processor.parsers.xlsx import XlsxParser


def _pricing_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Pricing"
    ws["A1"] = "Unit Price"
    ws["B1"] = 12.5
    ws["B1"].number_format = '"$"#,##0.00'
    ws["C1"] = "=B1*2"
    ws.merge_cells("A2:B2")
    ws["A2"] = "Merged labor category"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_xlsx_parser_reads_sheets_formulas_merged_cells_and_number_formats() -> None:
    parsed = XlsxParser().parse(_pricing_workbook(), mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="pricing.xlsx")
    assert parsed.parser_id == "xlsx-openpyxl"
    assert parsed.sheet_count == 1
    sheet = parsed.sheets[0]
    assert sheet.name == "Pricing"
    assert "A2:B2" in sheet.merged_ranges
    by_coord = {cell.coordinate: cell for cell in sheet.cells}
    assert "B2" not in by_coord
    assert by_coord["A2"].merged_range == "A2:B2"
    assert by_coord["B1"].number_format is not None
    assert by_coord["C1"].formula == "=B1*2"
    assert by_coord["C1"].cached_value in {"25", "25.0", None} or (
        by_coord["C1"].cached_value is not None and "25" in by_coord["C1"].cached_value
    )


def test_heuristic_extractor_emits_cell_facts_without_model() -> None:
    parsed = XlsxParser().parse(_pricing_workbook(), mime_type=None, filename="pricing.xlsx")
    drafts = HeuristicExtractor().extract(parsed)
    keys = [draft.idempotency_key for draft in drafts]
    assert len(keys) == len(set(keys))
    fields = {draft.field for draft in drafts}
    assert "Pricing!B1" in fields
    assert "Pricing!C1" in fields
